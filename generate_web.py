"""
generate_web.py — Genera los archivos de salida del scraper.

Salidas:
  1. vinyls_YYYY-MM-DD.xlsx  — Excel multi-hoja
  2. web/data/vinyls.json    — JSON compacto para GitHub Pages

Excel multi-hoja:
  - "Productos"     : todos los productos con campos completos
  - "Resumen"       : conteo por tienda + % disponibilidad
  - "Comparación"   : delta vs run anterior por tienda
  - "Errores"       : tiendas con problemas del run actual

JSON compacto para la web:
  Campos: {a, al, p, av, l, s} = artist, album, price, available, link, store
  Solo disponibles se marcan con av=1; el resto av=0
"""

import json
import logging
from datetime import date
from pathlib import Path
from typing import Optional

from scrapers.base import Product, ScrapeError

logger = logging.getLogger(__name__)

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
WEB_DATA_DIR = ROOT / "web" / "data"
WEB_DATA_DIR.mkdir(parents=True, exist_ok=True)


# ─── Excel ────────────────────────────────────────────────────────────────────

def generate_excel(
    products: list[Product],
    errors: list[ScrapeError],
    sanity_alerts: list[str],
    results_by_store: dict[str, int],
    last_stats: dict,
    run_date: str,
) -> Path:
    """Genera el Excel multi-hoja. Retorna la ruta al archivo."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl no instalado. Instalar con: pip install openpyxl")
        return Path()

    wb = openpyxl.Workbook()

    # ── Hoja 1: Productos ─────────────────────────────────────────────────────
    ws_prod = wb.active
    ws_prod.title = "Productos"

    headers = ["Artista", "Álbum", "Precio (CLP)", "Disponible", "Tienda", "URL",
               "Artista Norm.", "Álbum Norm."]
    _write_header_row(ws_prod, headers)

    for p in sorted(products, key=lambda x: (x.store, x.artist.lower(), x.album.lower())):
        ws_prod.append([
            p.artist,
            p.album,
            p.price,
            "Sí" if p.available else "No",
            p.store,
            p.url,
            p.artist_norm or p.artist,
            p.album_norm or p.album,
        ])

    # Formato de precio como número
    for row in ws_prod.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'

    _auto_column_width(ws_prod)

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    _write_header_row(ws_sum, ["Tienda", "Total", "Disponibles", "% Disponible", "Precio Mín.", "Precio Máx.", "Precio Prom."])

    store_products: dict[str, list[Product]] = {}
    for p in products:
        store_products.setdefault(p.store, []).append(p)

    for store_name in sorted(store_products):
        plist = store_products[store_name]
        available = sum(1 for p in plist if p.available)
        prices = [p.price for p in plist if p.price > 0]
        ws_sum.append([
            store_name,
            len(plist),
            available,
            f"{available / len(plist):.0%}" if plist else "0%",
            min(prices) if prices else 0,
            max(prices) if prices else 0,
            int(sum(prices) / len(prices)) if prices else 0,
        ])

    _auto_column_width(ws_sum)

    # ── Hoja 3: Comparación con run anterior ─────────────────────────────────
    ws_cmp = wb.create_sheet("Comparación")
    _write_header_row(ws_cmp, ["Tienda", "Actual", "Anterior", "Delta", "% Cambio", "Estado"])

    prev_stores = last_stats.get("stores", {})
    for store_name in sorted(set(list(results_by_store.keys()) + list(prev_stores.keys()))):
        current = results_by_store.get(store_name, 0)
        previous = prev_stores.get(store_name, {}).get("count", 0)
        delta = current - previous
        pct = f"{delta / previous:.0%}" if previous > 0 else "N/A"
        if previous == 0:
            status = "NUEVO"
        elif current == 0:
            status = "SIN DATOS"
        elif delta < -previous * 0.5:
            status = "⚠️ CAÍDA >50%"
        elif delta > previous * 0.5:
            status = "↑ CRECIMIENTO"
        else:
            status = "OK"
        ws_cmp.append([store_name, current, previous, delta, pct, status])

    _auto_column_width(ws_cmp)

    # ── Hoja 4: Errores ───────────────────────────────────────────────────────
    ws_err = wb.create_sheet("Errores")
    _write_header_row(ws_err, ["Tienda", "Tipo Error", "Mensaje", "Recuperable"])

    if errors:
        for err in errors:
            ws_err.append([
                err.store_name,
                err.error_type,
                err.message[:200],
                "Sí" if err.recoverable else "No",
            ])
    else:
        ws_err.append(["(Sin errores en este run)", "", "", ""])

    if sanity_alerts:
        ws_err.append([])
        ws_err.append(["⚠️ ALERTAS SANITY CHECK", "", "", ""])
        for alert in sanity_alerts:
            ws_err.append([alert, "", "", ""])

    _auto_column_width(ws_err)

    # Guardar
    filename = f"vinyls_{run_date}.xlsx"
    path = ROOT / filename
    wb.save(path)
    size_mb = path.stat().st_size / (1024 * 1024)
    logger.info(f"Excel generado: {path} ({size_mb:.1f} MB, {len(products)} productos)")
    return path


def _write_header_row(ws, headers: list[str]) -> None:
    """Escribe fila de encabezado con estilo."""
    try:
        from openpyxl.styles import Font, PatternFill
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
    except ImportError:
        ws.append(headers)


def _auto_column_width(ws, max_width: int = 60) -> None:
    """Ajusta ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


# ─── JSON para la web ─────────────────────────────────────────────────────────

def generate_json(products: list[Product]) -> Path:
    """
    Genera web/data/vinyls.json con campos compactos:
      {a, al, p, av, l, s} = artist, album, price, available, link, store

    También genera web/data/meta.json con metadatos del run.
    """
    data = []
    for p in products:
        data.append({
            "a": p.artist_norm or p.artist,
            "al": p.album_norm or p.album,
            "p": p.price,
            "av": 1 if p.available else 0,
            "l": p.url,
            "s": p.store,
        })

    # Ordenar por tienda, luego artista
    data.sort(key=lambda x: (x["s"], x["a"].lower(), x["al"].lower()))

    json_path = WEB_DATA_DIR / "vinyls.json"
    json_path.write_text(
        json.dumps(data, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    size_kb = json_path.stat().st_size / 1024
    logger.info(f"JSON web generado: {json_path} ({size_kb:.0f} KB, {len(data)} productos)")

    # Metadatos del run
    stores_summary = {}
    for item in data:
        s = item["s"]
        if s not in stores_summary:
            stores_summary[s] = {"total": 0, "available": 0}
        stores_summary[s]["total"] += 1
        if item["av"]:
            stores_summary[s]["available"] += 1

    meta = {
        "run_date": date.today().isoformat(),
        "total": len(data),
        "available": sum(1 for d in data if d["av"]),
        "stores": stores_summary,
    }
    meta_path = WEB_DATA_DIR / "meta.json"
    meta_path.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    return json_path
